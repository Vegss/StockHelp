import openpyxl as xl
import services.sheets as sheetService
from datetime import datetime
import os
import sys
import json

def createWorkbook(stockSymbols):
	if len(stockSymbols) == 0:
		return None
	if os.path.exists("updated.json"):
		with open("updated.json", "r") as f:
			info = json.load(f)
			if info["updated"] == datetime.now().strftime("%Y-%m-%d") and info["stocks"] == stockSymbols:
				print("Already updated today")
				os.system('powershell.exe /C ./stocks.xlsx')
				sys.exit()

	workbook = xl.Workbook()
	workbook.remove(workbook.active)

	for symbol in stockSymbols:
			sheetService.createSheet(symbol, workbook)

	workbook.save("stocks.xlsx")

	info = {
		"updated": datetime.now().strftime("%Y-%m-%d"),
		"stocks": stockSymbols
	}

	with open("updated.json", "w") as outfile:
		json.dump(info, outfile)

if __name__ == "__main__":
	stockSymbols = ["AAPL", "MSFT", "AMZN", "GOOG", "TSLA", "NFLX", "NVDA", "PYPL", "ADBE"]
	createWorkbook(stockSymbols)
	os.system('powershell.exe /C ./stocks.xlsx')